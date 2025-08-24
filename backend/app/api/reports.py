# Reports API routes
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, date, timedelta
from ..core.database import get_db_cursor
from ..core.utils import create_response, require_external_auth, validate_required_fields, log_activity
import pandas as pd
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/attendance', methods=['GET'])
@require_external_auth
def get_attendance_report(current_user_id):
    """Tạo báo cáo chấm công"""
    try:
        # Lấy parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        user_id = request.args.get('user_id')
        format_type = request.args.get('format', 'json')  # json, excel, csv
        
        # Kiểm tra quyền admin nếu xem báo cáo của user khác
        with get_db_cursor() as cursor:
            cursor.execute("SELECT role FROM users WHERE id = %s", (current_user_id,))
            user_role = cursor.fetchone()
            
            if user_id and user_id != str(current_user_id):
                if not user_role or user_role[0] != 'admin':
                    return create_response(False, error='Admin access required', status_code=403)
            
            # Xây dựng query
            base_query = """
                SELECT a.id, a.date, a.check_in_time, a.check_out_time, 
                       a.total_hours, a.status, u.username, u.full_name
                FROM attendance a
                JOIN users u ON a.user_id = u.id
                WHERE 1=1
            """
            params = []
            
            if user_id:
                base_query += " AND a.user_id = %s"
                params.append(user_id)
            elif user_role and user_role[0] != 'admin':
                # Non-admin users can only see their own data
                base_query += " AND a.user_id = %s"
                params.append(current_user_id)
            
            if start_date:
                base_query += " AND a.date >= %s"
                params.append(start_date)
            
            if end_date:
                base_query += " AND a.date <= %s"
                params.append(end_date)
            
            base_query += " ORDER BY a.date DESC, u.full_name"
            
            cursor.execute(base_query, params)
            records = cursor.fetchall()
            
            # Tính toán thống kê
            total_days = len(records)
            total_hours = sum(float(record[4] or 0) for record in records)
            present_days = len([r for r in records if r[5] == 'present'])
            late_days = len([r for r in records if r[5] == 'late'])
            absent_days = len([r for r in records if r[5] == 'absent'])
            
            report_data = {
                'summary': {
                    'total_days': total_days,
                    'total_hours': round(total_hours, 2),
                    'present_days': present_days,
                    'late_days': late_days,
                    'absent_days': absent_days,
                    'attendance_rate': round((present_days + late_days) / total_days * 100, 2) if total_days > 0 else 0
                },
                'records': [
                    {
                        'id': record[0],
                        'date': record[1].strftime('%Y-%m-%d'),
                        'check_in_time': record[2].strftime('%H:%M:%S') if record[2] else None,
                        'check_out_time': record[3].strftime('%H:%M:%S') if record[3] else None,
                        'total_hours': float(record[4]) if record[4] else 0,
                        'status': record[5],
                        'username': record[6],
                        'full_name': record[7]
                    }
                    for record in records
                ],
                'generated_at': datetime.now().isoformat(),
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                }
            }
            
            if format_type == 'json':
                return create_response(True, report_data)
            
            elif format_type == 'excel':
                return generate_excel_report(report_data, 'attendance_report')
            
            elif format_type == 'csv':
                return generate_csv_report(report_data['records'], 'attendance_report')
            
            else:
                return create_response(False, error='Invalid format type', status_code=400)
                
    except Exception as e:
        log_activity('ERROR', f'Attendance report error: {str(e)}', 'reports')
        return create_response(False, error=f'Report generation failed: {str(e)}', status_code=500)

@reports_bp.route('/salary', methods=['GET'])
@require_external_auth
def get_salary_report(current_user_id):
    """Tạo báo cáo lương"""
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        user_id = request.args.get('user_id')
        format_type = request.args.get('format', 'json')
        
        if not month or not year:
            return create_response(False, error='Month and year are required', status_code=400)
        
        with get_db_cursor() as cursor:
            cursor.execute("SELECT role FROM users WHERE id = %s", (current_user_id,))
            user_role = cursor.fetchone()
            
            if user_id and user_id != str(current_user_id):
                if not user_role or user_role[0] != 'admin':
                    return create_response(False, error='Admin access required', status_code=403)
            
            # Lấy dữ liệu chấm công cho tháng
            query = """
                SELECT u.id, u.username, u.full_name, u.email,
                       COUNT(a.id) as total_days,
                       SUM(CASE WHEN a.total_hours IS NOT NULL THEN a.total_hours ELSE 0 END) as total_hours,
                       COUNT(CASE WHEN a.status = 'present' THEN 1 END) as present_days,
                       COUNT(CASE WHEN a.status = 'late' THEN 1 END) as late_days,
                       COUNT(CASE WHEN a.status = 'absent' THEN 1 END) as absent_days
                FROM users u
                LEFT JOIN attendance a ON u.id = a.user_id 
                    AND EXTRACT(MONTH FROM a.date) = %s 
                    AND EXTRACT(YEAR FROM a.date) = %s
                WHERE u.role = 'user'
            """
            params = [month, year]
            
            if user_id:
                query += " AND u.id = %s"
                params.append(user_id)
            elif user_role and user_role[0] != 'admin':
                query += " AND u.id = %s"
                params.append(current_user_id)
            
            query += " GROUP BY u.id, u.username, u.full_name, u.email ORDER BY u.full_name"
            
            cursor.execute(query, params)
            users_data = cursor.fetchall()
            
            # Tính lương (công thức cơ bản)
            base_salary = 5000000  # 5 triệu VND cơ bản
            hourly_rate = 50000   # 50k/giờ
            
            salary_data = []
            for user in users_data:
                total_hours = float(user[5] or 0)
                present_days = user[6]
                late_days = user[7]
                
                # Tính lương cơ bản theo ngày công
                working_days = present_days + late_days
                daily_salary = base_salary / 22  # 22 ngày làm việc/tháng
                base_amount = working_days * daily_salary
                
                # Thưởng giờ làm thêm
                overtime_hours = max(0, total_hours - (working_days * 8))
                overtime_amount = overtime_hours * hourly_rate
                
                # Phạt đi trễ
                late_penalty = late_days * 50000  # 50k/lần trễ
                
                total_salary = base_amount + overtime_amount - late_penalty
                
                salary_data.append({
                    'user_id': user[0],
                    'username': user[1],
                    'full_name': user[2],
                    'email': user[3],
                    'working_days': working_days,
                    'total_hours': total_hours,
                    'present_days': present_days,
                    'late_days': late_days,
                    'base_amount': round(base_amount, 0),
                    'overtime_hours': round(overtime_hours, 2),
                    'overtime_amount': round(overtime_amount, 0),
                    'late_penalty': round(late_penalty, 0),
                    'total_salary': round(total_salary, 0)
                })
            
            report_data = {
                'period': f"{month:02d}/{year}",
                'generated_at': datetime.now().isoformat(),
                'salary_data': salary_data,
                'summary': {
                    'total_employees': len(salary_data),
                    'total_payroll': sum(s['total_salary'] for s in salary_data)
                }
            }
            
            if format_type == 'json':
                return create_response(True, report_data)
            elif format_type == 'excel':
                return generate_excel_salary_report(report_data)
            else:
                return create_response(False, error='Invalid format type', status_code=400)
                
    except Exception as e:
        log_activity('ERROR', f'Salary report error: {str(e)}', 'reports')
        return create_response(False, error=f'Salary report generation failed: {str(e)}', status_code=500)

@reports_bp.route('/dashboard', methods=['GET'])
@require_external_auth
def get_dashboard_stats(current_user_id):
    """Lấy thống kê cho dashboard"""
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT role FROM users WHERE id = %s", (current_user_id,))
            user_role = cursor.fetchone()
            is_admin = user_role and user_role[0] == 'admin'
            
            today = date.today()
            this_month_start = today.replace(day=1)
            
            if is_admin:
                # Admin dashboard - toàn bộ hệ thống
                stats = {}
                
                # Tổng số người dùng
                cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'user'")
                stats['total_users'] = cursor.fetchone()[0]
                
                # Chấm công hôm nay
                cursor.execute("""
                    SELECT COUNT(DISTINCT user_id) FROM attendance 
                    WHERE date = %s
                """, (today,))
                stats['today_checkins'] = cursor.fetchone()[0]
                
                # Tổng giờ làm tháng này
                cursor.execute("""
                    SELECT COALESCE(SUM(total_hours), 0) FROM attendance 
                    WHERE date >= %s
                """, (this_month_start,))
                stats['month_total_hours'] = float(cursor.fetchone()[0] or 0)
                
                # Đơn xin nghỉ chờ duyệt
                cursor.execute("""
                    SELECT COUNT(*) FROM leave_requests 
                    WHERE status = 'pending'
                """)
                stats['pending_leave_requests'] = cursor.fetchone()[0]
                
                # Top 5 nhân viên chăm chỉ tháng này
                cursor.execute("""
                    SELECT u.full_name, COALESCE(SUM(a.total_hours), 0) as total_hours
                    FROM users u
                    LEFT JOIN attendance a ON u.id = a.user_id AND a.date >= %s
                    WHERE u.role = 'user'
                    GROUP BY u.id, u.full_name
                    ORDER BY total_hours DESC
                    LIMIT 5
                """, (this_month_start,))
                stats['top_employees'] = [
                    {'name': row[0], 'hours': float(row[1])}
                    for row in cursor.fetchall()
                ]
                
            else:
                # User dashboard - chỉ dữ liệu cá nhân
                stats = {}
                
                # Giờ làm hôm nay
                cursor.execute("""
                    SELECT total_hours FROM attendance 
                    WHERE user_id = %s AND date = %s
                """, (current_user_id, today))
                result = cursor.fetchone()
                stats['today_hours'] = float(result[0]) if result and result[0] else 0
                
                # Giờ làm tuần này
                week_start = today - timedelta(days=today.weekday())
                cursor.execute("""
                    SELECT COALESCE(SUM(total_hours), 0) FROM attendance 
                    WHERE user_id = %s AND date >= %s
                """, (current_user_id, week_start))
                stats['week_hours'] = float(cursor.fetchone()[0] or 0)
                
                # Giờ làm tháng này
                cursor.execute("""
                    SELECT COALESCE(SUM(total_hours), 0) FROM attendance 
                    WHERE user_id = %s AND date >= %s
                """, (current_user_id, this_month_start))
                stats['month_hours'] = float(cursor.fetchone()[0] or 0)
                
                # Số ngày đi làm tháng này
                cursor.execute("""
                    SELECT COUNT(*) FROM attendance 
                    WHERE user_id = %s AND date >= %s AND status IN ('present', 'late')
                """, (current_user_id, this_month_start))
                stats['month_working_days'] = cursor.fetchone()[0]
                
                # Đơn xin nghỉ của user
                cursor.execute("""
                    SELECT status, COUNT(*) FROM leave_requests 
                    WHERE user_id = %s 
                    GROUP BY status
                """, (current_user_id,))
                leave_stats = dict(cursor.fetchall())
                stats['leave_requests'] = {
                    'pending': leave_stats.get('pending', 0),
                    'approved': leave_stats.get('approved', 0),
                    'rejected': leave_stats.get('rejected', 0)
                }
            
            return create_response(True, {
                'stats': stats,
                'user_role': user_role[0] if user_role else 'user',
                'generated_at': datetime.now().isoformat()
            })
            
    except Exception as e:
        log_activity('ERROR', f'Dashboard stats error: {str(e)}', 'reports')
        return create_response(False, error=f'Failed to get dashboard stats: {str(e)}', status_code=500)

def generate_excel_report(data, filename):
    """Tạo file Excel từ dữ liệu báo cáo"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Header
    headers = ['Date', 'Full Name', 'Check In', 'Check Out', 'Total Hours', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, col=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, record in enumerate(data['records'], 2):
        ws.cell(row=row, col=1, value=record['date'])
        ws.cell(row=row, col=2, value=record['full_name'])
        ws.cell(row=row, col=3, value=record['check_in_time'])
        ws.cell(row=row, col=4, value=record['check_out_time'])
        ws.cell(row=row, col=5, value=record['total_hours'])
        ws.cell(row=row, col=6, value=record['status'])
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

def generate_csv_report(records, filename):
    """Tạo file CSV từ dữ liệu báo cáo"""
    df = pd.DataFrame(records)
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'{filename}_{datetime.now().strftime("%Y%m%d")}.csv'
    )

def generate_excel_salary_report(data):
    """Tạo file Excel báo cáo lương"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Report"
    
    # Header
    headers = ['Full Name', 'Working Days', 'Total Hours', 'Base Amount', 
               'Overtime Hours', 'Overtime Amount', 'Late Penalty', 'Total Salary']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, col=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, record in enumerate(data['salary_data'], 2):
        ws.cell(row=row, col=1, value=record['full_name'])
        ws.cell(row=row, col=2, value=record['working_days'])
        ws.cell(row=row, col=3, value=record['total_hours'])
        ws.cell(row=row, col=4, value=record['base_amount'])
        ws.cell(row=row, col=5, value=record['overtime_hours'])
        ws.cell(row=row, col=6, value=record['overtime_amount'])
        ws.cell(row=row, col=7, value=record['late_penalty'])
        ws.cell(row=row, col=8, value=record['total_salary'])
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'salary_report_{data["period"].replace("/", "_")}.xlsx'
    )
